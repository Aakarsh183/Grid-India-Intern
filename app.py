import pandas as pd
import glob
import os
from openpyxl import load_workbook
from openpyxl.styles import  PatternFill,Font,Border,Side

folder_path = "C:/My workspace/Python/Data Sanity report"
excel_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
empty_frame = pd.DataFrame()
empty_frame.to_excel("Data Sanity Report(AAKARSH GERA).xlsx")
station_names = ['OSTRO','BHUVAD','VADVA','NARANPAR','DAYAPUR','BARANDA','SRIJAN','SITEC','POWERICA','MYTRAH','ORANGE','GIREL','BETAM WIND','JSWRETwo','AYANA SIX']
for file in excel_files:
    xls = pd.ExcelFile(file)
    index = 1
    empty_df = pd.DataFrame()
    lt = []
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls,sheet_name=sheet_name)
        i = 1
        while i < len(df.columns):
            k = i
            dictionary = {}
            dictionary['SNo.'] = index
            dictionary['Pooling station'] = sheet_name
            s = df.columns[k]
            idx = s.rindex('_')
            Station_name = s[:idx]
            dictionary['Station Name'] = Station_name
            if Station_name not in station_names:
                while k < i + 4:
                   s1 = df.columns[k]
                   idx1 = s1.rindex('_')
                   Parameter = s1[idx1+1:]
                   if Parameter == "INSTALL CAP.":
                       Parameter = "AVC"
                   dictionary[Parameter] = ""
                   result_list_missing_data = []
                   result_list_zero = []
                   result_list_negative = []
                   result_list_out_of_range = []
                   result_list_fixed = []
                   # Check for missing data
                   d = 3
                   while d < len(df):
                       value = df.iloc[d,k]
                       if file[42:50] != "WR_ISTS_":
                           if pd.isna(value) and 25<=d<=81:
                               start = d
                               while d + 1 < len(df) and pd.isna(df.iloc[d+1,k]) and 25 <= d + 1 <= 81:
                                   d += 1
                               end = d
                               if start == end:
                                   result_list_missing_data.append(f"at {str(df.iloc[start, 0])[11:16]}")
                               else:
                                   result_list_missing_data.append(f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")
                       else:
                           if pd.isna(value) and 26<=d<=82:
                               start = d
                               while d + 1 < len(df) and pd.isna(df.iloc[d+1,k]) and 26 <= d + 1 <= 82:
                                   d += 1
                               end = d
                               if start == end:
                                   result_list_missing_data.append(f"at {str(df.iloc[start, 0])[11:16]}")
                               else:
                                   result_list_missing_data.append(f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")

                       d += 1
                   # Check for zero data
                   t = 3
                   while t < len(df):
                      if file[42:50] != "WR_ISTS_":
                          if df.iloc[t,k] == 0 and 25 <= t <= 81:
                              start = t
                              while t + 1 < len(df) and df.iloc[t + 1, k] == 0 and 25 <= t + 1 <= 81:
                                  t += 1
                              end = t
                              if start == end:
                                  result_list_zero.append(f"at {str(df.iloc[start, 0])[11:16]}")
                              else:
                                  result_list_zero.append(f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")
                      else:
                          if df.iloc[t,k] == 0 and 26 <= t <= 82:
                              start = t
                              while t + 1 < len(df) and df.iloc[t + 1, k] == 0 and 26 <= t + 1 <= 82:
                                  t += 1
                              end = t
                              if start == end:
                                  result_list_zero.append(f"at {str(df.iloc[start, 0])[11:16]}")
                              else:
                                  result_list_zero.append(f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")
                      t += 1
                   # Check for negative data
                   n = 3
                   while n < len(df):
                       if file[42:50] != "WR_ISTS_":
                          if df.iloc[n,k] < 0 and 25 <= n <= 81:
                              start = n
                              while n + 1 < len(df) and df.iloc[n + 1, k] < 0 and 25 <= n + 1 <= 81:
                                  n += 1
                              end = n
                              if start == end:
                                  result_list_negative.append(f"at {str(df.iloc[start, 0])[11:16]}")
                              else:
                                  result_list_negative.append(f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")
                       else:
                           if df.iloc[n, k] < 0 and 26 <= n <= 82:
                               start = n
                               while n + 1 < len(df) and df.iloc[n + 1, k] < 0 and 26 <= n + 1 <= 82:
                                   n += 1
                               end = n
                               if start == end:
                                   result_list_negative.append(f"at {str(df.iloc[start, 0])[11:16]}")
                               else:
                                   result_list_negative.append(
                                       f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")
                       n += 1
                   # Check for data out of acceptable range
                   m = 3
                   while m < len(df):
                       if file[42:50] != "WR_ISTS_":
                           if pd.isna(df.loc[m,s]) or pd.isna(df.iloc[m,k]):
                               m += 1
                               continue
                           elif df.iloc[m,k] > df.loc[m,s] and df.iloc[m,k] > 0 and 25 <= m <= 81:
                               start = m
                               while m + 1 < len(df) and df.iloc[m + 1, k] > df.loc[m,s] and df.iloc[
                                   m + 1, k] > 0 and 25 <= m + 1 <= 81:
                                   m += 1
                               end = m
                               out_of_range = str(df.iloc[start,0])
                               if start == end:
                                   result_list_out_of_range.append(f"at {out_of_range[11:16]}")
                               else:
                                   result_list_out_of_range.append(f"from {str(df.iloc[start,0])[11:16]} to {str(df.iloc[end,0])[11:16]}")
                       else:
                           if pd.isna(df.loc[m,s]) or pd.isna(df.iloc[m,k]):
                               m += 1
                               continue
                           elif df.iloc[m,k] > df.loc[m,s] and df.iloc[m,k] > 0 and 26 <= m <= 82:
                               start = m
                               while m + 1 < len(df) and df.iloc[m + 1, k] > df.loc[m,s] and df.iloc[
                                   m + 1, k] > 0 and 26 <= m + 1 <= 82:
                                   m += 1
                               end = m
                               out_of_range = str(df.iloc[start,0])
                               if start == end:
                                   result_list_out_of_range.append(f"at {out_of_range[11:16]}")
                               else:
                                   result_list_out_of_range.append(f"from {str(df.iloc[start,0])[11:16]} to {str(df.iloc[end,0])[11:16]}")
                       m += 1
                   # Check for fixed data
                   j = 3
                   while j < len(df) - 1 and Parameter != "AVC":
                       a = j
                       if file[42:50] != "WR_ISTS_":
                           if (df.iloc[a,k] != 0) and (df.iloc[a,k] > 0) and (df.iloc[a,k] == df.iloc[a+1,k]) and (25 <= a <= 81):
                              start = str(df.iloc[a,0])
                              while a + 1 < len(df) and df.iloc[a,k] == df.iloc[a+1,k] and (25 <= a <= 81):
                                  a = a + 1
                              end = str(df.iloc[a,0])
                              result_list_fixed.append(f"from {start[11:16]} to {end[11:16]}")
                              j = a + 1
                           else:
                              j += 1
                       else:
                           if (df.iloc[a, k] != 0) and (df.iloc[a, k] > 0) and (df.iloc[a, k] == df.iloc[a + 1, k]) and (
                                   26 <= a <= 83):
                               start = str(df.iloc[a, 0])
                               while a + 1 < len(df) and df.iloc[a, k] == df.iloc[a + 1, k] and (26 <= a <= 82):
                                   a = a + 1
                               end = str(df.iloc[a, 0])
                               result_list_fixed.append(f"from {start[11:16]} to {end[11:16]}")
                               j = a + 1
                           else:
                               j += 1
                   if result_list_missing_data:
                       dictionary[Parameter] = "Data is missing " + " and ".join(result_list_missing_data)
                   elif result_list_zero:
                       dictionary[Parameter] = "Data is zero " + " and ".join(result_list_zero)
                   elif result_list_negative:
                       dictionary[Parameter] = "Data is negative " + " and ".join(result_list_negative)
                   elif result_list_out_of_range:
                       dictionary[Parameter] = "Data is out of range " + " and ".join(result_list_out_of_range)
                   elif result_list_fixed:
                       dictionary[Parameter] = "Data is fixed " + " and ".join(result_list_fixed)
                   else:
                       dictionary[Parameter] = "Data is of good quality"
                   k += 1
                i = i + 4
                index += 1
                lt.append(dictionary)
            else:
                while k < i + 4:
                    s1 = df.columns[k]
                    idx1 = s1.rindex('_')
                    Parameter = s1[idx1 + 1:]
                    dictionary[Parameter] = ""
                    result_list_missing_data = []
                    result_list_zero = []
                    result_list_negative = []
                    result_list_out_of_range = []
                    result_list_fixed = []
                    # Check for missing data
                    d = 3
                    while d < len(df):
                        value = df.iloc[d, k]
                        if file[42:50] != "WR_ISTS_":
                            if pd.isna(value):
                                start = d
                                while d + 1 < len(df) and pd.isna(df.iloc[d + 1, k]):
                                    d += 1
                                end = d
                                if start == end:
                                    result_list_missing_data.append(f"at {str(df.iloc[start, 0])[11:16]}")
                                else:
                                    result_list_missing_data.append(
                                        f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")
                        else:
                            if pd.isna(value) and d != 3:
                                start = d
                                while d + 1 < len(df) and pd.isna(df.iloc[d + 1, k]):
                                    d += 1
                                end = d
                                if start == end:
                                    result_list_missing_data.append(f"at {str(df.iloc[start, 0])[11:16]}")
                                else:
                                    result_list_missing_data.append(
                                        f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")

                        d += 1
                    # Check for zero data
                    t = 3
                    while t < len(df):
                        if file[42:50] != "WR_ISTS_":
                            if df.iloc[t, k] == 0:
                                start = t
                                while t + 1 < len(df) and df.iloc[t + 1, k] == 0:
                                    t += 1
                                end = t
                                if start == end:
                                    result_list_zero.append(f"at {str(df.iloc[start, 0])[11:16]}")
                                else:
                                    result_list_zero.append(
                                        f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")
                        else:
                            if df.iloc[t, k] == 0 and d!=3:
                                start = t
                                while t + 1 < len(df) and df.iloc[t + 1, k] == 0:
                                    t += 1
                                end = t
                                if start == end:
                                    result_list_zero.append(f"at {str(df.iloc[start, 0])[11:16]}")
                                else:
                                    result_list_zero.append(
                                        f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")
                        t += 1
                    # Check for negative data
                    n = 3
                    while n < len(df):
                        if file[42:50] != "WR_ISTS_":
                            if df.iloc[n, k] < 0:
                                start = n
                                while n + 1 < len(df) and df.iloc[n + 1, k] < 0:
                                    n += 1
                                end = n
                                if start == end:
                                    result_list_negative.append(f"at {str(df.iloc[start, 0])[11:16]}")
                                else:
                                    result_list_negative.append(
                                        f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")
                        else:
                            if df.iloc[n, k] < 0:
                                start = n
                                while n + 1 < len(df) and df.iloc[n + 1, k] < 0:
                                    n += 1
                                end = n
                                if start == end:
                                    result_list_negative.append(f"at {str(df.iloc[start, 0])[11:16]}")
                                else:
                                    result_list_negative.append(
                                        f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")
                        n += 1
                    # Check for data out of acceptable range
                    m = 3
                    while m < len(df):
                        if file[42:50] != "WR_ISTS_":
                            if pd.isna(df.loc[m, s]) or pd.isna(df.iloc[m, k]):
                                m += 1
                                continue
                            elif df.iloc[m, k] > df.loc[m, s] and df.iloc[m, k] > 0:
                                start = m
                                while m + 1 < len(df) and df.iloc[m + 1, k] > df.loc[m, s] and df.iloc[
                                    m + 1, k] > 0:
                                    m += 1
                                end = m
                                out_of_range = str(df.iloc[start, 0])
                                if start == end:
                                    result_list_out_of_range.append(f"at {out_of_range[11:16]}")
                                else:
                                    result_list_out_of_range.append(
                                        f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")
                        else:
                            if pd.isna(df.loc[m, s]) or pd.isna(df.iloc[m, k]) and d!=3:
                                m += 1
                                continue
                            elif df.iloc[m, k] > df.loc[m, s] and df.iloc[m, k] > 0:
                                start = m
                                while m + 1 < len(df) and df.iloc[m + 1, k] > df.loc[m, s] and df.iloc[
                                    m + 1, k] > 0:
                                    m += 1
                                end = m
                                out_of_range = str(df.iloc[start, 0])
                                if start == end:
                                    result_list_out_of_range.append(f"at {out_of_range[11:16]}")
                                else:
                                    result_list_out_of_range.append(
                                        f"from {str(df.iloc[start, 0])[11:16]} to {str(df.iloc[end, 0])[11:16]}")
                        m += 1
                    # Check for fixed data
                    j = 3
                    while j < len(df) - 1 and Parameter != "AVC":
                        a = j
                        if file[42:50] != "WR_ISTS_":
                            if (df.iloc[a, k] != 0) and (df.iloc[a, k] > 0) and (
                                    df.iloc[a, k] == df.iloc[a + 1, k]):
                                start = str(df.iloc[a, 0])
                                while a + 1 < len(df) and df.iloc[a, k] == df.iloc[a + 1, k]:
                                    a = a + 1
                                end = str(df.iloc[a, 0])
                                result_list_fixed.append(f"from {start[11:16]} to {end[11:16]}")
                                j = a + 1
                            else:
                                j += 1
                        else:
                            if (df.iloc[a, k] != 0) and (df.iloc[a, k] > 0) and (
                                    df.iloc[a, k] == df.iloc[a + 1, k] and d!=3):
                                start = str(df.iloc[a, 0])
                                while a + 1 < len(df) and df.iloc[a, k] == df.iloc[a + 1, k]:
                                    a = a + 1
                                end = str(df.iloc[a, 0])
                                result_list_fixed.append(f"from {start[11:16]} to {end[11:16]}")
                                j = a + 1
                            else:
                                j += 1
                    if result_list_missing_data:
                        dictionary[Parameter] = "Data is missing " + " and ".join(result_list_missing_data)
                    elif result_list_zero:
                        dictionary[Parameter] = "Data is zero " + " and ".join(result_list_zero)
                    elif result_list_negative:
                        dictionary[Parameter] = "Data is negative " + " and ".join(result_list_negative)
                    elif result_list_out_of_range:
                        dictionary[Parameter] = "Data is out of range " + " and ".join(result_list_out_of_range)
                    elif result_list_fixed:
                        dictionary[Parameter] = "Data is fixed " + " and ".join(result_list_fixed)
                    else:
                        dictionary[Parameter] = "Data is of good quality"
                    k += 1
                i = i + 4
                index += 1
                lt.append(dictionary)
    dataframe = pd.concat([empty_df, pd.DataFrame(lt)], ignore_index=True)
    with pd.ExcelWriter("Data Sanity Report(AAKARSH GERA).xlsx",engine="openpyxl",mode='a') as writer:
        dataframe.to_excel(writer, sheet_name=file[42:50], index=False)

    wb = load_workbook("Data Sanity Report(AAKARSH GERA).xlsx")
    ws = wb[file[42:50]] # Data in specified sheet
    for box in ws[1]:
        box.fill = PatternFill(start_color= "C6EFCE",end_color= "C6EFCE",fill_type= "solid")
    red_fill = PatternFill(start_color="FF0000",end_color="FF0000",fill_type="solid")
    blue_fill = PatternFill(start_color= "C6EFCE",end_color= "C6EFCE",fill_type= "solid")
    yellow_fill = PatternFill(start_color="FFEB9C",end_color="FFEB9C",fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9",end_color="D9D9D9",fill_type="solid")
    green_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
    lavender_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

    thick_border = Border(
        left=Side(border_style="thick", color="000000"),
        right=Side(border_style="thick", color="000000"),
        top=Side(border_style="thick", color="000000"),
        bottom=Side(border_style="thick", color="000000")
    )

    headers = {cell.column_letter: cell.value for cell in ws[1]}
    count1 = [0,0,0,0]
    count2 = [0,0,0,0]
    count3 = [0,0,0,0]
    count4 = [0,0,0,0]
    count5 = [0,0,0,0]
    count6 = [0,0,0,0]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thick_border
            if isinstance(cell.value, str):
                text = cell.value.lower()
                col_name = headers.get(cell.column_letter, f"Col {cell.column_letter}")
                if "data is fixed" in text:
                    cell.fill = red_fill
                    if col_name == "INSTALL CAP." or col_name == "AVC":
                       count1[0] += 1
                    if col_name == "FORECAST":
                       count1[1] += 1
                    if col_name == "SCHDEULE" or col_name == "SCHEDULE":
                       count1[2] += 1
                    if col_name == "ACTUAL":
                       count1[3] += 1
                elif "data is negative" in text:
                    cell.fill = blue_fill
                    if col_name == "INSTALL CAP." or col_name == "AVC":
                       count2[0] += 1
                    if col_name == "FORECAST":
                       count2[1] += 1
                    if col_name == "SCHDEULE" or col_name == "SCHEDULE":
                       count2[2] += 1
                    if col_name == "ACTUAL":
                       count2[3] += 1
                elif "data is out of range" in text:
                    cell.fill = yellow_fill
                    if col_name == "INSTALL_CAP." or col_name == "AVC":
                       count3[0] += 1
                    if col_name == "FORECAST":
                       count3[1] += 1
                    if col_name == "SCHDEULE" or col_name == "SCHEDULE":
                       count3[2] += 1
                    if col_name == "ACTUAL":
                       count3[3] += 1
                elif "data is zero" in text:
                    cell.fill = lavender_fill
                    if col_name == "INSTALL_CAP." or col_name == "AVC":
                       count4[0] += 1
                    if col_name == "FORECAST":
                       count4[1] += 1
                    if col_name == "SCHDEULE" or col_name == "SCHEDULE":
                       count4[2] += 1
                    if col_name == "ACTUAL":
                       count4[3] += 1
                elif "data is of good quality" in text:
                    cell.fill = green_fill
                    if col_name == "INSTALL_CAP." or col_name == "AVC":
                       count5[0] += 1
                    if col_name == "FORECAST":
                       count5[1] += 1
                    if col_name == "SCHDEULE" or col_name == "SCHEDULE":
                       count5[2] += 1
                    if col_name == "ACTUAL":
                       count5[3] += 1
                elif "data is missing" in text:
                    cell.fill = gray_fill
                    if col_name == "INSTALL CAP." or col_name == "AVC":
                        count6[0] += 1
                    if col_name == "FORECAST":
                        count6[1] += 1
                    if col_name == "SCHDEULE" or col_name == "SCHEDULE":
                        count6[2] += 1
                    if col_name == "ACTUAL":
                        count6[3] += 1
    start_row = ws.max_row + 3  # 3 rows below current data
    start_col = 1
    s_row = ws.max_row + 12
    s_col = 1
    # Legend entries and colors
    legend = [
        ("Data is updating with good quality and it is within the range.", "A9D08E"),  # light green
        ("Data is constant (unchanging)", "FF0000"),               # red
        ("Data is zero.", "E6E6FA"),                                # lavender
        ("Data is out of acceptable range.", "FFEB9C"),                               # pale yellow
        ("Data is negative (logically invalid).", "C6EFCE"),                          # blue
        ("Data is not available / missing.", "D9D9D9")                                # gray
    ]

    legend_summary = [
        ("Data is updating with good quality and it is within the range.", f"{count5[0]}/{ws.max_row-1}",f"{count5[1]}/{ws.max_row-1}",f"{count5[2]}/{ws.max_row-1}",f"{count5[3]}/{ws.max_row-1}"),  # light green
        ("Data is constant (unchanging)", f"{count1[0]}/{ws.max_row-1}",f"{count1[1]}/{ws.max_row-1}",f"{count1[2]}/{ws.max_row-1}",f"{count1[3]}/{ws.max_row-1}"),               # light red
        ("Data is zero.",f"{count4[0]}/{ws.max_row-1}",f"{count4[1]}/{ws.max_row-1}",f"{count4[2]}/{ws.max_row-1}",f"{count4[3]}/{ws.max_row-1}"),                                # pale yellow                                  # light red
        ("Data is out of acceptable range.",f"{count3[0]}/{ws.max_row-1}",f"{count3[1]}/{ws.max_row-1}",f"{count3[2]}/{ws.max_row-1}",f"{count3[3]}/{ws.max_row-1}"),                               # orange
        ("Data is negative (logically invalid).",f"{count2[0]}/{ws.max_row-1}",f"{count2[1]}/{ws.max_row-1}",f"{count2[2]}/{ws.max_row-1}",f"{count2[3]}/{ws.max_row-1}"),                          # lavender
        ("Data is not available / missing.",f"{count6[0]}/{ws.max_row-1}",f"{count6[1]}/{ws.max_row-1}",f"{count6[2]}/{ws.max_row-1}",f"{count6[3]}/{ws.max_row-1}")                                # gray
    ]

    # Header row
    ws.cell(row=start_row, column=start_col, value="Legend Description").font = Font(bold=True)
    ws.cell(row=start_row, column=start_col, value="Legend Description").border = thick_border
    ws.cell(row=start_row, column=start_col + 1, value="Color").font = Font(bold=True)
    ws.cell(row=start_row, column=start_col + 1, value="Color").border = thick_border

    ws.cell(row=s_row-1,column=s_col+3,value="SUMMARY REPORT").font = Font(bold=True)

    ws.cell(row=s_row, column=s_col, value="Flag Type").font = Font(bold=True)
    ws.cell(row=s_row, column=s_col, value="Flag Type").fill = blue_fill
    ws.cell(row=s_row, column=s_col, value="Flag Type").border = thick_border
    ws.cell(row=s_row, column=s_col + 1, value="AVC").font = Font(bold=True)
    ws.cell(row=s_row, column=s_col + 1, value="AVC").fill = blue_fill
    ws.cell(row=s_row, column=s_col + 1, value="AVC").border = thick_border
    ws.cell(row=s_row, column=s_col + 2, value="FORECAST").font = Font(bold=True)
    ws.cell(row=s_row, column=s_col + 2, value="FORECAST").fill = blue_fill
    ws.cell(row=s_row, column=s_col + 2, value="FORECAST").border = thick_border
    ws.cell(row=s_row, column=s_col + 3, value="SCHDEULE").font = Font(bold=True)
    ws.cell(row=s_row, column=s_col + 3, value="SCHDEULE").fill = blue_fill
    ws.cell(row=s_row, column=s_col + 3, value="SCHDEULE").border = thick_border
    ws.cell(row=s_row, column=s_col + 4, value="ACTUAL").font = Font(bold=True)
    ws.cell(row=s_row, column=s_col + 4, value="ACTUAL").fill = blue_fill
    ws.cell(row=s_row, column=s_col + 4, value="ACTUAL").border = thick_border

    for idx, (desc, color_hex) in enumerate(legend, start=1):
        r = start_row + idx
        ws.cell(row=r, column=start_col, value=desc).border = thick_border
        color_cell = ws.cell(row=r, column=start_col + 1, value="")
        color_cell.border = thick_border
        color_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")


    var = 0
    colour_fill = ["A9D08E","FF0000","E6E6FA","FFEB9C","C6EFCE","D9D9D9"]
    for idx, (desc,v1,v2,v3,v4) in enumerate(legend_summary, start=1):
        r = s_row + idx
        ws.cell(row=r, column=s_col, value=desc).fill = PatternFill(start_color = colour_fill[var],end_color= colour_fill[var],fill_type= "solid")
        ws.cell(row=r, column=s_col, value=desc).border = thick_border
        ws.cell(row=r, column=s_col + 1, value=v1).fill = PatternFill(start_color= colour_fill[var],end_color= colour_fill[var],fill_type= "solid")
        ws.cell(row=r, column=s_col + 1, value=v1).border = thick_border
        ws.cell(row=r, column=s_col + 2, value=v2).fill = PatternFill(start_color= colour_fill[var],end_color= colour_fill[var],fill_type= "solid")
        ws.cell(row=r, column=s_col + 2, value=v2).border = thick_border
        ws.cell(row=r, column=s_col + 3, value=v3).fill = PatternFill(start_color= colour_fill[var],end_color= colour_fill[var],fill_type= "solid")
        ws.cell(row=r, column=s_col + 3, value=v3).border = thick_border
        ws.cell(row=r, column=s_col + 4, value=v4).fill = PatternFill(start_color= colour_fill[var],end_color= colour_fill[var],fill_type= "solid")
        ws.cell(row=r, column=s_col + 4, value=v4).border = thick_border
        var += 1

    wb.save("Data Sanity Report(AAKARSH GERA).xlsx")
output = load_workbook("Data Sanity Report(AAKARSH GERA).xlsx")
for sheet in output.sheetnames:
    if sheet == 'Sheet1':
      del output[sheet]
output.save("Data Sanity Report(AAKARSH GERA).xlsx")





























